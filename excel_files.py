import os

import pandas as pd

def combine_excel_files_to_single_df(path=None):
    """
    read all excel in folder and combines them into one df
    """
    df = pd.DataFrame()
    for dirName, subdirs, fileList in os.walk(path):
        for filename in fileList:
             path = os.path.join(dirName, filename)
             print(f"Searching file: {path}...")
             df = df.append(pd.read_excel(path), ignore_index=True, sort=False)
    return df

def read_excel_with_pandas(filename):
    df = pd.read_excel(filename)
    print(df)
    # for index, row in df.tail(5).iterrows():
    #     print(index, row)
    #     if "Total" in row:
    #         print(f"found row at {index}")

def read_excel_with_xlrd(filename):
    import xlrd
    # Open the workbook
    xl_workbook = xlrd.open_workbook(filename)
    # List sheet names, and pull a sheet by name
    sheet_names = xl_workbook.sheet_names()
    print('Sheet Names', sheet_names)
    sheet = xl_workbook.sheet_by_index(0)
    # Extracting number of rows
    print(sheet.nrows)
    for row in range(sheet.nrows): # Iterates over your sheet
        row_value = sheet.row_values(row)
        print(row_value)

def read_excel_with_openpyxl(filename, sheet_name=None):
    # https://github.com/anishst/Learn/blob/master/Programming/Python/excel_related/ReadExcel_openpyxl.py
    from openpyxl import load_workbook
    wb = load_workbook(filename=filename)
    if sheet_name:
        # print(wb.get_sheet_names())
        ws = wb.get_sheet_by_name(sheet_name)

    # +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    # Loop excel file: reads all but separate lines [ FINAL ]
    # +++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
    for row in ws.iter_rows():
        for cell in row:
            print(cell.value, sep="\n")

filename = "DataFiles/xls_format.xls"
read_excel_with_pandas(filename)
# read_excel_with_openpyxl(filename)
read_excel_with_xlrd(filename)